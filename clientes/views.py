from django.utils.html import escape
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.utils.translation import gettext as _

from .models import Cliente, Direccion, ImportacionLog, AgenteVentas
from .forms import ClienteForm, DireccionForm, DireccionFormSet, ImportacionForm

#from django.shortcuts import render

### Exportar a Excel ####
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

### Exportar a Pdf ####
from django.template.loader import get_template
from xhtml2pdf import pisa

### Importación planilla de clientes ####
import hashlib
import json
from openpyxl import load_workbook


# Helper para chequear si el usuario es supervisor
def is_supervisor(user):
    return user.groups.filter(name='Supervisor').exists()

@login_required
def lista_clientes(request):
    if is_supervisor(request.user):
        clientes = Cliente.objects.all()
    else:
        # filtro por agente relacionado al usuario
        clientes = Cliente.objects.filter(agente__user=request.user)
    return render(request, 'clientes/lista.html', {
        'clientes': clientes, 
        'es_supervisor': is_supervisor(request.user)
    })

@login_required
def crear_cliente(request):
    """
    Crear cliente:
    - Agentes: no pueden elegir el agente, se les asigna automáticamente.
    - Supervisores: pueden elegir cualquier agente en el formulario.
    """
    # Instanciamos el form
    cliente_form = ClienteForm(request.POST or None)
    # Si es agente, removemos el campo 'agente' del form para que no lo vea
    if not is_supervisor(request.user) and 'agente' in cliente_form.fields:
        del cliente_form.fields['agente']

    direccion_formset = DireccionFormSet(request.POST or None)

    if request.method == 'POST':
        if cliente_form.is_valid() and direccion_formset.is_valid():
  
            # Guardamos cliente sin commit para asignar agente si es necesario
            cliente = cliente_form.save(commit=False)

            if not is_supervisor(request.user):
                try:
                    agente = AgenteVentas.objects.get(user=request.user)
                    cliente.agente = agente
                except AgenteVentas.DoesNotExist:
                    messages.error(request, "Tu usuario no está registrado como agente de ventas.")
                    return redirect('lista_clientes')
            else:
                cliente.agente = cliente_form.cleaned_data['agente']

            cliente.save()
            # Asociamos y guardamos el formset de direcciones
            direccion_formset = DireccionFormSet(request.POST, instance=cliente)
            if direccion_formset.is_valid():
                direccion_formset.save()
                messages.success(request, _("Cliente y direcciones guardados correctamente."))
                return redirect('lista_clientes')

    return render(request, 'clientes/formulario.html', {
        'form': cliente_form,
        'direccion_formset': direccion_formset,
    })

@login_required
def editar_cliente(request, pk):
    """
    Edita los datos básicos de un cliente.
    Las direcciones se manejan en sus propias vistas de crear/editar/eliminar.
    """
    cliente = get_object_or_404(Cliente, pk=pk)

    # Control de permisos: solo supervisor o el agente propietario
    if not is_supervisor(request.user) and cliente.agente.user != request.user:
        return HttpResponseForbidden("No tienes permiso para editar este cliente.")

    if request.method == "POST":
        form = ClienteForm(request.POST, instance=cliente)
        # Si no es supervisor, impedir que cambie el agente
        if not is_supervisor(request.user) and 'agente' in form.fields:
            del form.fields['agente']

        if form.is_valid():
            form.save()
            messages.success(request, _("Cliente actualizado correctamente."))
            return redirect("lista_clientes")
    else:
        form = ClienteForm(instance=cliente)
        if not is_supervisor(request.user) and 'agente' in form.fields:
            del form.fields['agente']

    # Obtiene las direcciones existentes para mostrar en la plantilla
    direcciones = cliente.direcciones.all()

    return render(request, "clientes/formulario.html", {
        "form": form,
        "cliente": cliente,
        "direcciones": direcciones,
        # Pasamos un flag para que la plantilla siga su lógica original
        "direccion_formset": True,
    })


@login_required
def confirmar_eliminar_cliente(request, pk):
    """
    Confirmar y eliminar cliente:
    - Sólo propietario o supervisor.
    """
    cliente = get_object_or_404(Cliente, pk=pk)
    if not is_supervisor(request.user) and cliente.agente.user != request.user:
        return HttpResponseForbidden("No tienes permiso para eliminar este cliente.")

    if request.method == 'POST':
        cliente.delete()
        messages.success(request, f'El cliente "{cliente.nombre_razon_social}" fue eliminado correctamente.')
        return redirect('lista_clientes')

    return render(request, 'clientes/confirmar_eliminar_cliente.html', {'cliente': cliente})

@login_required
def agregar_direccion(request, cliente_id):
    """
    Agregar dirección a un cliente existente:
    - Sólo propietario o supervisor.
    """
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    
    if not (is_supervisor(request.user) or cliente.agente.user == request.user):
        return HttpResponseForbidden(
            f"No tienes permiso para agregar direcciones a este cliente. "
            f"(Agente asignado: {escape(str(cliente.agente.user))} | Usuario actual: {escape(str(request.user))})"
        )

    form = DireccionForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        direccion = form.save(commit=False)
        direccion.cliente = cliente
        direccion.save()
        messages.success(request, _("Dirección agregada correctamente."))
        return redirect('editar_cliente', pk=cliente.id)

    return render(request, 'clientes/form_direccion.html', {
        'form': form,
        'cliente': cliente,
        'modo': 'agregar'
    })

@login_required
def editar_direccion(request, pk):
    """
    Editar una dirección:
    - Sólo propietario o supervisor.
    """
    direccion = get_object_or_404(Direccion, pk=pk)
    cliente = direccion.cliente
    if not is_supervisor(request.user) and cliente.agente.user != request.user:
        return HttpResponseForbidden("No tienes permiso para editar esta dirección.")

    form = DireccionForm(request.POST or None, instance=direccion)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, _("Dirección actualizada correctamente."))
        return redirect('editar_cliente', pk=cliente.id)

    return render(request, 'clientes/form_direccion.html', {
        'form': form,
        'cliente': cliente,
        'modo': 'editar'
    })

@login_required
def eliminar_direccion(request, pk):
    """
    Eliminar una dirección:
    - Sólo propietario o supervisor.
    """
    direccion = get_object_or_404(Direccion, pk=pk)
    cliente = direccion.cliente
    if not is_supervisor(request.user) and cliente.agente.user != request.user:
        # print(type(cliente.agente), cliente.agente.user)
        # print(type(request.user), request.user)
        return HttpResponseForbidden("No tienes permiso para eliminar esta dirección.")
    

    if request.method == 'POST':
        direccion.delete()
        messages.success(request, _("La dirección fue eliminada correctamente."))
        return redirect('editar_cliente', pk=cliente.pk)

    return render(request, 'clientes/confirmar_eliminar_direccion.html', {
        'direccion': direccion,
        'cliente': cliente
    })


############################### Vista de Login/Logout #####################
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('lista_clientes')
        else:
            error = "Usuario o contraseña inválidos"
    else:
        error = None

    return render(request, 'login.html', { 'error': error })

def logout_view(request):
    logout(request)
    return redirect('login')


######################### Vistas de Consulta y Exportación a Excel/Pdf #####################
@login_required
def consulta_clientes(request):
    clientes = Cliente.objects.select_related('agente').prefetch_related('direcciones')
    return render(request, 'clientes/consulta.html', {
        'clientes': clientes
    })

#====================================
# Vista para exportar a Excel
#====================================
def exportar_clientes_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes y Direcciones"

    # Encabezados
    ws.append(["Cliente", "Correo", "Comuna", "Ciudad", "Dirección", "País"])

    # Datos
    clientes = Cliente.objects.prefetch_related('direcciones')
    for cliente in clientes:
        for direccion in cliente.direcciones.all():
            ws.append([
                cliente.nombre_razon_social,
                cliente.email,
                direccion.comuna,
                direccion.ciudad,
                f"{direccion.calle} {direccion.numero}",
                direccion.pais
            ])

    # Ajustar tamaño de columnas
    for col in ws.columns:
        max_length = 0
        columna = col[0].column
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ajusta = get_column_letter(columna)
        ws.column_dimensions[ajusta].width = max_length + 2

    # Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes_direcciones.xlsx'
    wb.save(response)
    return response

#====================================
# Vista para exportar a Pdf
#====================================
def exportar_clientes_pdf(request):
    clientes = Cliente.objects.prefetch_related('direcciones')
    template = get_template('clientes/pdf_clientes.html')
    html = template.render({'clientes': clientes})

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=clientes_direcciones.pdf'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse("Error al generar el PDF", status=500)
    return response


#=========================================
# Vista para importar planilla de clientes
#=========================================
@login_required
#@user_passes_test(es_supervisor)
def importar_clientes(request):
    """
    Vista para subir un Excel y crear clientes + direcciones.
    Muestra errores inline y solo redirige tras éxito real.
    """
    if request.method == 'POST':
        form = ImportacionForm(request.POST, request.FILES)

        # 1) ¿Llega archivo?
        archivo = request.FILES.get('archivo')
        if not archivo:
            form.add_error('archivo', 'Debes seleccionar un archivo antes de importar.')
        # 2) ¿Extensión válida?
        elif not archivo.name.lower().endswith(('.xlsx', '.xls')):
            form.add_error('archivo', 'Formato no válido. Usa .xlsx o .xls.')

        # 3) Si el form está libre de errores de campo, seguimos con hash y duplicados
        if not form.errors:
            # Leer contenido para hash
            contenido = archivo.read()
            hash_archivo = hashlib.sha256(contenido).hexdigest()

            # Reinicia puntero para usar luego en openpyxl
            archivo.seek(0)

            # 4) ¿Ya existe ese hash?
            if ImportacionLog.objects.filter(hash_archivo=hash_archivo).exists():
                form.add_error(None, 'Este archivo ya fue importado anteriormente.')

        # 5) Si tras todas las validaciones el form está OK → procesar y redirigir
        if form.is_valid():
            # Guardar log preliminar
            log = form.save(commit=False)
            log.usuario = request.user
            log.hash_archivo = hash_archivo
            log.save()

            # Cargar el libro
            wb = load_workbook(archivo, data_only=True)
            sheet = wb.active
            fila_inicio = 2

            errores = []
            exitosos = fallidos = 0

            for idx, row in enumerate(sheet.iter_rows(min_row=fila_inicio), start=fila_inicio):
                tipo, nombre, rut, email, telefono, web, obs_cli, \
                  calle, numero, comuna, ciudad, cp, pais, obs_dir = (
                    cell.value for cell in row
                )

                # Validaciones básicas
                if not (nombre and rut and comuna and calle):
                    fallidos += 1
                    errores.append({'fila': idx, 'error': 'Faltan campos obligatorios'})
                    continue

                if Cliente.objects.filter(rut=rut).exists():
                    fallidos += 1
                    errores.append({'fila': idx, 'error': 'Cliente ya existe'})
                    continue

                # Crear Cliente + Dirección
                cliente = Cliente.objects.create(
                    tipo_entidad=tipo,
                    nombre_razon_social=nombre,
                    rut=rut,
                    email=email,
                    telefono=telefono,
                    sitio_web=web,
                    observacion=obs_cli
                )
                Direccion.objects.create(
                    cliente=cliente,
                    calle=calle,
                    numero=numero,
                    comuna=comuna,
                    ciudad=ciudad,
                    codigo_postal=cp,
                    pais=pais,
                    observacion=obs_dir
                )
                exitosos += 1

            # Actualizar log con resultados
            log.exitosos = exitosos
            log.fallidos = fallidos
            log.errores = json.dumps(errores, ensure_ascii=False, indent=2)
            log.save()

            # Mensaje de éxito y redirect (Post/Redirect/Get)
            messages.success(
                request,
                f"Importación finalizada: {exitosos} clientes creados, {fallidos} errores."
            )
            return redirect('lista_clientes')  # Asume que tu vista de lista tiene este name

        # Si hay ANY error de form (campo o no-field), caemos aquí
        # Y volvemos a renderizar la vista con los errores inline
        return render(request, 'clientes/importar_clientes.html', {'form': form})

    # GET: primera carga
    form = ImportacionForm()
    return render(request, 'clientes/importar_clientes.html', {'form': form})


######################### Vista para el Dashboard #####################
@login_required
#@user_passes_test(es_supervisor)
def dashboard_supervisor(request):
    total_clientes = Cliente.objects.count()
    sin_agente = Cliente.objects.filter(agente__isnull=True).count()
    activos = Cliente.objects.filter(activo=True).count()
    inactivos = Cliente.objects.filter(activo=False).count()
    ultima_importacion = ImportacionLog.objects.order_by('-fecha').first()

    agentes = AgenteVentas.objects.all()
    clientes_por_agente = {
        agente.nombre: Cliente.objects.filter(agente=agente).count()
        for agente in agentes
    }

    context = {
        'total_clientes': total_clientes,
        'sin_agente': sin_agente,
        'activos': activos,
        'inactivos': inactivos,
        'ultima_importacion': ultima_importacion,
        'clientes_por_agente': clientes_por_agente,
    }
    return render(request, 'clientes/dashboard.html', context)
